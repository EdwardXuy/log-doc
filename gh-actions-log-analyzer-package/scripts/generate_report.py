#!/usr/bin/env python3
"""GitHub Actions Log Analyzer - Report Generator"""
import csv, os, argparse
from datetime import datetime
from collections import defaultdict

def pct(n, d): return round(n/d*100,1) if d>0 else 0
def safe(t, m=40):
    t = t.replace("|","/")
    return t[:m]+"..." if len(t)>m else t

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--timestamp", required=True)
    parser.add_argument("--base-dir", default=".")
    parser.add_argument("--repo", default="sgl-project/sgl-kernel-npu")
    args = parser.parse_args()
    base = os.path.abspath(args.base_dir)
    rdir = os.path.join(base, "analysis", "reports", args.timestamp)
    odir = os.path.join(base, "analysis", "optimization", args.timestamp)
    os.makedirs(odir, exist_ok=True)
    with open(os.path.join(rdir,"all_runs.csv"), encoding="utf-8-sig") as f: runs=list(csv.DictReader(f))
    with open(os.path.join(rdir,"all_jobs.csv"), encoding="utf-8-sig") as f: jobs=list(csv.DictReader(f))
    dt=datetime.now().strftime("%Y-%m-%d %H:%M:%S"); repo=args.repo
    wr_d=defaultdict(list); wj_d=defaultdict(list)
    for r in runs: wr_d[r["WorkflowName"]].append(r)
    for j in jobs: wj_d[j["WorkflowName"]].append(j)

    L=[]
    L.append(f"# {repo} CI/CD 流水线分析报告")
    L.append(""); L.append(f"**生成时间**: {dt}"); L.append(f"**仓库**: {repo}")
    L.append("**分析范围**: 所有已完成的运行（成功 + 失败）")
    L.append(f"**数据量**: {len(runs)} 次运行, {len(jobs)} 个任务"); L.append(""); L.append("---"); L.append("")

    L.append("## 1. 概览"); L.append("")
    for wf in ["build_and_release","daily-build-test","pr-test-npu"]:
        wr=wr_d.get(wf,[]); wj=wj_d.get(wf,[])
        if not wr: continue
        rs=sum(1 for r in wr if r["Conclusion"]=="success"); rf=len(wr)-rs
        js=sum(1 for j in wj if j["JobConclusion"]=="success"); jf=sum(1 for j in wj if j["JobConclusion"]=="failure")
        jc=sum(1 for j in wj if j["JobConclusion"]=="cancelled"); jsk=sum(1 for j in wj if j["JobConclusion"]=="skipped")
        eff=len(wj)-jc-jsk
        L.append(f"### {wf}"); L.append("")
        L.append("| 指标 | 数值 |"); L.append("|------|------|")
        L.append(f"| 总运行次数 | {len(wr)} |"); L.append(f"| 成功 / 失败 | {rs} / {rf} |")
        L.append(f"| 运行成功率 | {pct(rs,len(wr))}% |"); L.append(f"| 总任务数 | {len(wj)} |")
        L.append(f"| 成功 / 失败 / 取消 / 跳过 | {js} / {jf} / {jc} / {jsk} |")
        L.append(f"| 任务通过率（不含取消/跳过） | {pct(js,eff)}% |"); L.append("")

    trs=sum(1 for r in runs if r["Conclusion"]=="success")
    tjs=sum(1 for j in jobs if j["JobConclusion"]=="success"); tjf=sum(1 for j in jobs if j["JobConclusion"]=="failure")
    tjc=sum(1 for j in jobs if j["JobConclusion"]=="cancelled"); tjsk=sum(1 for j in jobs if j["JobConclusion"]=="skipped")
    teff=len(jobs)-tjc-tjsk
    L.append("### 总计"); L.append("")
    L.append("| 指标 | 数值 |"); L.append("|------|------|")
    L.append(f"| 总运行次数 | {len(runs)} |"); L.append(f"| 成功 / 失败 | {trs} / {len(runs)-trs} |")
    L.append(f"| 运行成功率 | {pct(trs,len(runs))}% |"); L.append(f"| 总任务数 | {len(jobs)} |")
    L.append(f"| 成功 / 失败 / 取消 / 跳过 | {tjs} / {tjf} / {tjc} / {tjsk} |")
    L.append(f"| 任务通过率（不含取消/跳过） | {pct(tjs,teff)}% |"); L.append("")
    L.append("---"); L.append("")

    L.append("## 2. 每次运行的任务统计（最近20次）"); L.append("")
    for wf in ["build_and_release","daily-build-test","pr-test-npu"]:
        wr=wr_d.get(wf,[])
        if not wr: continue
        L.append(f"### {wf}"); L.append("")
        L.append("| 运行 | 标题 | 结论 | 总任务 | 成功 | 失败 | 取消 | 通过率 |")
        L.append("|------|------|------|--------|------|------|------|--------|")
        for r in sorted(wr, key=lambda r:r["CreatedAt"],reverse=True)[:20]:
            rj=[j for j in wj_d.get(wf,[]) if j["RunId"]==r["RunId"]]
            s=sum(1 for j in rj if j["JobConclusion"]=="success"); f=sum(1 for j in rj if j["JobConclusion"]=="failure")
            c=sum(1 for j in rj if j["JobConclusion"]=="cancelled"); eff=len(rj)-c; rate=pct(s,eff)
            mark="成功" if r["Conclusion"]=="success" else "失败"; title=safe(r.get("Title","")); url=r.get("HtmlUrl","")
            L.append(f"| [{r['RunId']}]({url}) | {title} | {mark} | {len(rj)} | {s} | {f} | {c} | {rate}% |")
        L.append("")
    L.append("---"); L.append("")

    L.append("## 3. 稳定性分析"); L.append("")
    for wf in ["pr-test-npu","daily-build-test"]:
        wj=[j for j in wj_d.get(wf,[]) if j["JobConclusion"]!="cancelled"]
        if not wj: continue
        L.append(f"### {wf} 任务稳定性"); L.append("")
        L.append("| 任务名称 | 总次数 | 成功 | 失败 | 通过率 |"); L.append("|----------|--------|------|------|--------|")
        js={}
        for j in wj:
            n=j["JobName"]
            if n not in js: js[n]={"total":0,"success":0,"failure":0}
            js[n]["total"]+=1
            if j["JobConclusion"]=="success": js[n]["success"]+=1
            elif j["JobConclusion"]=="failure": js[n]["failure"]+=1
        for n in sorted(js, key=lambda x:js[x]["total"],reverse=True):
            st=js[n]; L.append(f"| {n} | {st['total']} | {st['success']} | {st['failure']} | {pct(st['success'],st['total'])}% |")
        L.append("")

    L.append("### PR 工作流端到端时长（目标: <= 60 分钟）"); L.append("")
    L.append("| 运行 | 标题 | 结论 | 核心任务最大时长 | 是否达标 |")
    L.append("|------|------|------|------------------|----------|")
    for r in sorted(wr_d.get("pr-test-npu",[]), key=lambda r:r["CreatedAt"],reverse=True)[:30]:
        rj=[j for j in wj_d.get("pr-test-npu",[]) if j["RunId"]==r["RunId"] and j["JobConclusion"]!="cancelled" and "Check changed files" not in j["JobName"] and j["JobName"]!="finish"]
        if not rj: continue
        durs=[float(j["DurationMin"]) for j in rj if float(j.get("DurationMin",0))>0]
        mx=round(max(durs),1) if durs else 0; meets="是" if mx<=60 else "否"
        title=safe(r.get("Title",""),30); url=r.get("HtmlUrl","")
        L.append(f"| [{r['RunId']}]({url}) | {title} | {r['Conclusion']} | {mx} 分钟 | {meets} |")
    L.append(""); L.append("---"); L.append("")

    L.append("## 4. 执行时间分析"); L.append("")
    tj=[j for j in jobs if j["JobConclusion"]!="cancelled" and float(j.get("DurationMin",0))>0]
    tj.sort(key=lambda j:float(j["DurationMin"]),reverse=True)
    L.append("### 4.1 耗时最长的30个任务（所有工作流）"); L.append("")
    L.append("| 排名 | 工作流 | 任务名称 | 运行ID | 时长(分钟) | 结论 |")
    L.append("|------|--------|----------|--------|------------|------|")
    for i,j in enumerate(tj[:30],1):
        L.append(f"| {i} | {j['WorkflowName']} | {j['JobName']} | {j['RunId']} | {float(j['DurationMin'])} | {j['JobConclusion']} |")
    L.append("")

    L.append("### 4.2 按任务名称的平均时长"); L.append("")
    L.append("| 工作流 | 任务名称 | 次数 | 平均(分钟) | 最大(分钟) | 最小(分钟) |")
    L.append("|--------|----------|------|------------|------------|------------|")
    ds={}
    for j in tj:
        k=(j["WorkflowName"],j["JobName"])
        if k not in ds: ds[k]=[]
        ds[k].append(float(j["DurationMin"]))
    for (wf,nm),durs in sorted(ds.items(), key=lambda x:-max(x[1])):
        L.append(f"| {wf} | {nm} | {len(durs)} | {round(sum(durs)/len(durs),1)} | {round(max(durs),1)} | {round(min(durs),1)} |")
    L.append(""); L.append("---"); L.append("")

    L.append("## 5. 失败分析"); L.append("")
    fj=[j for j in jobs if j["JobConclusion"]=="failure"]
    L.append("### 5.1 按工作流统计失败任务"); L.append("")
    L.append("| 工作流 | 失败任务数 | 有效任务总数 | 失败率 |"); L.append("|--------|------------|--------------|--------|")
    for wf in ["build_and_release","daily-build-test","pr-test-npu"]:
        wf_f=[j for j in fj if j["WorkflowName"]==wf]; wf_e=[j for j in wj_d.get(wf,[]) if j["JobConclusion"]!="cancelled"]
        L.append(f"| {wf} | {len(wf_f)} | {len(wf_e)} | {pct(len(wf_f),len(wf_e))}% |")
    L.append("")

    L.append("### 5.2 按任务名称统计失败"); L.append("")
    L.append("| 工作流 | 任务名称 | 失败次数 | 常见失败步骤 |"); L.append("|--------|----------|----------|--------------|")
    fbn={}
    for j in fj:
        k=(j["WorkflowName"],j["JobName"])
        if k not in fbn: fbn[k]={"count":0,"steps":set()}
        fbn[k]["count"]+=1
        for s in j.get("FailedStepsNames","").split("; "):
            if s: fbn[k]["steps"].add(s)
    for (wf,nm),info in sorted(fbn.items(), key=lambda x:-x[1]["count"]):
        ss=", ".join(sorted(info["steps"]))[:80]
        L.append(f"| {wf} | {nm} | {info['count']} | {ss} |")
    L.append(""); L.append("---"); L.append("")
    L.append(f"*报告由 GitHub Actions 日志分析器生成 | 数据源: {repo}*")

    with open(os.path.join(rdir,"full_analysis_report.md"),"w",encoding="utf-8") as f: f.write("\n".join(L))
    print(f"Report saved: {os.path.join(rdir,'full_analysis_report.md')}")

    # Optimization report
    O=[]
    O.append(f"# {repo} CI/CD 流水线优化建议"); O.append("")
    O.append(f"生成时间: {dt}"); O.append("范围: 在不修改测试脚本的前提下减少流水线总执行时间"); O.append("")
    pt=[j for j in wj_d.get("pr-test-npu",[]) if j["JobConclusion"]!="cancelled" and "Check changed files" not in j["JobName"] and j["JobName"]!="finish" and float(j.get("DurationMin",0))>0]
    O.append("## 1. 当前 PR 工作流时间分解"); O.append("")
    O.append("| 任务名称 | 次数 | 平均(分钟) | 最大(分钟) |"); O.append("|----------|------|------------|------------|")
    pd={}
    for j in pt:
        n=j["JobName"]
        if n not in pd: pd[n]=[]
        pd[n].append(float(j["DurationMin"]))
    for n in sorted(pd, key=lambda n:-max(pd[n])):
        durs=pd[n]; O.append(f"| {n} | {len(durs)} | {round(sum(durs)/len(durs),1)} | {round(max(durs),1)} |")
    O.append("")
    O.append("## 2. 瓶颈分析"); O.append("")
    for n in sorted(pd, key=lambda n:-sum(pd[n])/len(pd[n])):
        durs=pd[n]; O.append(f"- **{n}**: 平均 {round(sum(durs)/len(durs),1)} 分钟, 最大 {round(max(durs),1)} 分钟")
    O.append("")
    O.append("## 3. 优化建议"); O.append("")
    O.append("### 建议 1: 将 test-all-build 拆分为并行子任务"); O.append("")
    O.append("现状: test-all-build 串行运行所有 DeepEP 测试（45+ 步骤, ~63 分钟）")
    O.append("建议: 按测试类型拆分为 3 个并行任务:"); O.append("")
    O.append("| 新任务 | 包含测试 | 预估时长 |"); O.append("|--------|----------|----------|")
    O.append("| test-intranode | test_intranode (13 种变体) | ~15 分钟 |")
    O.append("| test-low-latency-moe | test_low_latency (8) + test_fused_deep_moe (8) + test_mixed_running (7) | ~25 分钟 |")
    O.append("| test-combine-misc | test_combine (1) + test_generalization_fused_deep_moe (1) | ~10 分钟 |"); O.append("")
    O.append("预期改进: 墙上时间从 ~63 分钟降至 ~25 分钟（减少60%）"); O.append("")
    O.append("### 建议 2: 跨任务共享构建产物"); O.append("")
    O.append("现状: test-all-build, test-build-deepep-a3, test-build-deepep-a2 各自独立构建 DeepEP")
    O.append("建议: 创建专用构建任务，将 wheel 作为 GitHub Actions 产物上传")
    O.append("预期改进: 每次运行节省 20-30 分钟冗余构建时间"); O.append("")
    O.append("### 建议 3: 条件化 Internode 测试"); O.append("")
    O.append("现状: test_internode_a2 在每个 PR 都运行，失败率约90%，阻塞工作流 3 小时")
    O.append("建议: continue-on-error: true / 仅定时触发 / 超时从 10800秒 降至 1800秒")
    O.append("预期改进: K8s 不可用时 PR 工作流从 3+ 小时降至 <30 分钟"); O.append("")
    O.append("### 建议 4: 优化 enumerate_test 脚本"); O.append("")
    O.append("现状: enumerate 脚本对每个参数组合串行运行测试")
    O.append("建议: 使用后台进程或 GNU parallel 并行运行参数组合")
    O.append("预期改进: daily-build-test 任务时长减少 30-50%"); O.append("")
    O.append("### 建议 5: 将 sgl_kernel_npu 算子测试加入 CI"); O.append("")
    O.append("现状: tests/python/sgl_kernel_npu/ 中的 36 个算子测试未在任何 CI 工作流中")
    O.append("建议: 在 daily-build-test.yml 或 pr-test-npu.yml 中新增任务"); O.append("")
    O.append("## 4. 预期改进汇总"); O.append("")
    O.append("| 建议 | 目标 | 预期节省时间 | 优先级 |"); O.append("|------|------|--------------|--------|")
    O.append("| 拆分 test-all-build | PR 工作流 | ~38 分钟 (63 到 25 分钟) | P0 |")
    O.append("| 共享构建产物 | PR 工作流 | ~20 到 30 分钟 | P0 |")
    O.append("| 条件化 internode | PR 工作流 | ~2.5 小时 (K8s 故障时) | P0 |")
    O.append("| 并行 enumerate | 定时工作流 | 30% 到 50% 任务时间 | P1 |")
    O.append("| 添加算子测试 | 覆盖率 | N/A (新增测试) | P1 |"); O.append("")
    with open(os.path.join(odir,"optimization_proposal.md"),"w",encoding="utf-8") as f: f.write("\n".join(O))
    print(f"Optimization report saved: {os.path.join(odir,'optimization_proposal.md')}")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb=Workbook()
        hf=Font(bold=True,color="FFFFFF",size=11); hfl=PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
        tb=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        def sh(ws,mc):
            for c in range(1,mc+1):
                cl=ws.cell(row=1,column=c);cl.font=hf;cl.fill=hfl;cl.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True);cl.border=tb
        def aw(ws):
            for col in ws.columns:
                ml=0;cn=col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value))>ml:ml=len(str(cell.value))
                    except:pass
                ws.column_dimensions[cn].width=min(ml+4,60)

        ws1=wb.active;ws1.title="Per-Run Stats"
        h1=["Run ID","URL","Workflow","Title","Conclusion","Total Jobs","Success","Failure","Cancelled","Pass Rate"]
        for ci,h in enumerate(h1,1):ws1.cell(row=1,column=ci,value=h)
        sh(ws1,len(h1));ri=2
        for r in sorted(runs,key=lambda x:x["CreatedAt"],reverse=True):
            rj=[j for j in jobs if j["RunId"]==r["RunId"]]
            s=sum(1 for j in rj if j["JobConclusion"]=="success");f=sum(1 for j in rj if j["JobConclusion"]=="failure")
            c=sum(1 for j in rj if j["JobConclusion"]=="cancelled");eff=len(rj)-c;rate=f"{pct(s,eff)}%"
            vals=[r["RunId"],r.get("HtmlUrl",""),r["WorkflowName"],safe(r.get("Title",""),50),r["Conclusion"],len(rj),s,f,c,rate]
            for ci,v in enumerate(vals,1):cl=ws1.cell(row=ri,column=ci,value=v);cl.border=tb;cl.alignment=Alignment(vertical='center',wrap_text=True)
            ri+=1
        aw(ws1)

        ws2=wb.create_sheet("Job Duration")
        h2=["Workflow","Job Name","Count","Avg (min)","Max (min)","Min (min)","Success","Failure","Pass Rate"]
        for ci,h in enumerate(h2,1):ws2.cell(row=1,column=ci,value=h)
        sh(ws2,len(h2));ri=2
        for (wf,nm),durs in sorted(ds.items(),key=lambda x:-max(x[1])):
            avg=round(sum(durs)/len(durs),1);mx=round(max(durs),1);mn=round(min(durs),1)
            wf_j=[j for j in wj_d.get(wf,[]) if j["JobName"]==nm and j["JobConclusion"]!="cancelled"]
            s=sum(1 for j in wf_j if j["JobConclusion"]=="success");f=sum(1 for j in wf_j if j["JobConclusion"]=="failure")
            rate=f"{pct(s,len(wf_j))}%"
            vals=[wf,nm,len(durs),avg,mx,mn,s,f,rate]
            for ci,v in enumerate(vals,1):cl=ws2.cell(row=ri,column=ci,value=v);cl.border=tb;cl.alignment=Alignment(vertical='center',wrap_text=True)
            ri+=1
        aw(ws2)

        ws3=wb.create_sheet("Stability Summary")
        h3=["Workflow","Job Name","Total Runs","Success","Failure","Pass Rate"]
        for ci,h in enumerate(h3,1):ws3.cell(row=1,column=ci,value=h)
        sh(ws3,len(h3));ri=2
        for wf in ["pr-test-npu","daily-build-test","build_and_release"]:
            wj=[j for j in wj_d.get(wf,[]) if j["JobConclusion"]!="cancelled"];jss={}
            for j in wj:
                n=j["JobName"]
                if n not in jss:jss[n]={"total":0,"success":0,"failure":0}
                jss[n]["total"]+=1
                if j["JobConclusion"]=="success":jss[n]["success"]+=1
                elif j["JobConclusion"]=="failure":jss[n]["failure"]+=1
            for n in sorted(jss,key=lambda x:-jss[x]["total"]):
                st=jss[n];rate=f"{pct(st['success'],st['total'])}%"
                vals=[wf,n,st["total"],st["success"],st["failure"],rate]
                for ci,v in enumerate(vals,1):cl=ws3.cell(row=ri,column=ci,value=v);cl.border=tb;cl.alignment=Alignment(vertical='center',wrap_text=True)
                ri+=1
        aw(ws3)

        xp=os.path.join(rdir,"analysis_report.xlsx");wb.save(xp);print(f"Excel saved: {xp}")
    except ImportError:print("Warning: openpyxl not installed, skipping Excel")
    print("\n=== Report Generation Complete ===")

if __name__=="__main__":main()
